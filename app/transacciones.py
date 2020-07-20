# python 3.6
# Este script calcula las variaciones a partir del listado de transacciones, por lo que coincide exactamente con los
# resumenes de eToro, mientras que el de irpf solo considera operaciones cerradas.

from collections import defaultdict
from openpyxl import load_workbook

import elegir_fichero


if __name__ == '__main__':
    # Elijo el fichero de los disponibles en el directorio. Debe llamarse data-*
    file = elegir_fichero.menu()

    # cargo datos desde el excel
    workbook = load_workbook(filename=file)
    sheet = workbook['Transactions Report']

    # proceso los datos
    estructura = defaultdict(dict)
    total_profit_euros = 0
    total_fees_euros = 0
    cerradas = []
    for e in sheet.iter_rows(min_row=2,
                             max_col=9,
                             values_only=True):
        if e[2] == 'Profit/Loss of Trade':
            cerradas.append(e[4])
    equity_change_cerradas = 0
    equity_change_abiertas = 0
    fondos_aportados = 0
    for e in sheet.iter_rows(min_row=2,
                             max_col=9,
                             values_only=True):
        if e[4] in cerradas:
            equity_change_cerradas += e[6]
        else:
            if e[2] == 'Deposit':
                fondos_aportados += e[5]
            else:
                equity_change_abiertas += e[6]
    print('Fondos aportados en el período:                          ', '{:>8}'.format('{:.2f}'.format(fondos_aportados) + '$'))
    print()
    print('Equity change ordenes cerradas en el período:            ', '{:>8}'.format('{:.2f}'.format(equity_change_cerradas) + '$'))
    print('Equity change ordenes pendientes de cerrar en el período:', '{:>8}'.format('{:.2f}'.format(equity_change_abiertas) + '$'))
    print('Equity change en el período:                             ', '{:>8}'.format('{:.2f}'.format(equity_change_cerradas + equity_change_abiertas) + '$'))

