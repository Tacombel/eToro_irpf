# python 3.6

import os
from datetime import datetime, timedelta
from collections import defaultdict
import requests
import os.path
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from xml.etree.ElementTree import ParseError


def menu():
    path = './'
    counter = 1
    opciones = []
    # r=root, d=directories, f = files
    for f1 in os.walk(path):
        for file1 in f1[2]:
            if 'eToro' in file1:
                opciones.append(file1)
                print(counter, ': ', file1)
                counter += 1
        opcion_menu = input("Elije un fichero >> ")
        print('Procesaremos el fichero:', opciones[int(opcion_menu) - 1])
        return opciones[int(opcion_menu) - 1]


def rate_dolar(fecha2, fecha3):
    # Building blocks for the URL
    fecha2 = fecha2.strftime('%Y') + '-' + fecha2.strftime('%m') + '-' + fecha2.strftime('%d')
    fecha3 = fecha3.strftime('%Y') + '-' + fecha3.strftime('%m') + '-' + fecha3.strftime('%d')
    entrypoint = 'https://sdw-wsrest.ecb.europa.eu/service/'  # Using protocol 'https'
    resource = 'data'  # The resource for data queries is always'data'
    flowref = 'EXR'  # Dataflow describing the data that needs to be returned, exchange rates in this case
    key_ecb = 'D.USD.EUR.SP00.A'  # Defining the dimension values, explained below

    # Define the parameters
    parameters = {
        'startPeriod': fecha2,  # Start date of the time series
        'endPeriod': fecha3,  # End of the time series
    }
    # Construct the URL
    request_url = entrypoint + resource + '/' + flowref + '/' + key_ecb
    # Make the HTTP request
    response = requests.get(request_url, params=parameters)

    # Check if the response returns succesfully with response code 200
    # print(response)
    cambios = []
    data = response.text
    try:
        tree = ET.fromstring(data)
    except ParseError:
        print('No hay datos', response)
        return cambios
    ns = {'generic':"http://www.sdmx.org/resources/sdmxml/schemas/v2_1/data/generic"}
    dimensions = tree.findall('.//generic:Obs', namespaces=ns)
    for dim in dimensions:
        ObsDimension = dim.find('generic:ObsDimension', namespaces=ns)
        fecha = ObsDimension.get('value')
        ObsValue = dim.find('generic:ObsValue', namespaces=ns)
        # para hacer ahorrarme operaciones el que interesa es el $/EUR
        valor = 1/float(ObsValue.get('value'))
        print(fecha, valor)
        cambios.append([fecha, valor])
    if len(cambios) == 0:
        print('No hay datos nuevos')
    elif len(cambios) == 1:
        print('Hay un dato nuevo')
    elif len(cambios) > 1:
        print('Hay', len(cambios), 'datos nuevos')
    return cambios

def adaptar_fecha(fecha3):
    fecha_modificada = fecha3.split(" ")[0]
    fecha_modificada = fecha_modificada.split("/")
    fecha_modificada = fecha_modificada[2] + '-' + fecha_modificada[1] + '-' + fecha_modificada[0]
    return fecha_modificada


if __name__ == '__main__':
    # selecciono una fecha inicial para descargar cotizaciones del dolar
    fecha_inicial = '2019-12-31'
    fecha_inicial = datetime.strptime(fecha_inicial, "%Y-%m-%d")
    # Aqui almaceno los tipos de cambio que voy a necesitar. Para forzar la descarga, borrarlo
    cambio_euro_dolar = []
    cambio_euro_dolar_bruto = []
    if os.path.isfile('cambio_euro_dolar.txt'):
        with open('cambio_euro_dolar.txt', 'r') as filehandle:
            for line in filehandle:
                line = line[:-1]
                line = line[:-1]
                line = line[1:]
                line = line.split(',')
                line[0] = line[0].strip("'")
                line[1] = float(line[1])
                cambio_euro_dolar_bruto.append(line)
        for e in cambio_euro_dolar_bruto:
            cambio_euro_dolar.append([datetime.strptime(e[0], "%Y-%m-%d"), e[1]])
        cambio_euro_dolar = sorted(cambio_euro_dolar, reverse=True)

    today = datetime.today()
    today = datetime(today.year, today.month, today.day)
    print(today.strftime('Hoy es %d-%m-%Y'))

    if len(cambio_euro_dolar) == 0:
        ultima_fecha = fecha_inicial
    else:
        ultima_fecha = cambio_euro_dolar[0][0]

    print(ultima_fecha.strftime('Último cambio almacenado el %d-%m-%Y'))
    if today > ultima_fecha:
        nuevos_cambios = rate_dolar(ultima_fecha + timedelta(days=1), today)
        cambio_euro_dolar_bruto = cambio_euro_dolar_bruto + nuevos_cambios
        cambio_euro_dolar = []
        for e in cambio_euro_dolar_bruto:
            cambio_euro_dolar.append([datetime.strptime(e[0], "%Y-%m-%d"), e[1]])
        cambio_euro_dolar = sorted(cambio_euro_dolar, reverse=True)
        with open('cambio_euro_dolar.txt', 'w') as filehandle:
            for listitem in cambio_euro_dolar_bruto:
                listitem = str(listitem).strip('(')
                listitem = str(listitem).strip(')')
                filehandle.write('%s\n' % str(listitem))

    file = menu()

    # cargo datos desde el excel
    workbook = load_workbook(filename=file)
    sheet_1 = workbook['Closed Positions']

    # proceso los datos de posiciones cerradas
    estructura = defaultdict(dict)
    adquisiciones = 0
    num_adquisiciones = 0
    transmisiones = 0
    num_transmissiones = 0
    total_profit_euros = 0
    total_fees_euros = 0
    ID_operaciones_cerradas = []
    for e in sheet_1.iter_rows(min_row=2, max_col=15, values_only=True):
        # creo una lista de operaciones cerradas
        ID_operaciones_cerradas.append(e[0])
        # este bloque calcula el total en euros
        inicial_dolar = float(e[3].replace(',', '.'))
        # calculo el valor final de la operacion sumando el beneficio al valor inicial
        final_dolar = inicial_dolar + float(e[8].replace(',', '.'))
        inicial_fecha = datetime.strptime(adaptar_fecha(e[9]), "%Y-%m-%d")
        for cambio in cambio_euro_dolar:
            if inicial_fecha >= cambio[0]:
                inicial_cambio = cambio[1]
                break
        final_fecha = datetime.strptime(adaptar_fecha(e[10]), "%Y-%m-%d")
        for cambio in cambio_euro_dolar:
            if final_fecha >= cambio[0]:
                final_cambio = cambio[1]
                break
        inicial_euro = inicial_dolar * inicial_cambio
        adquisiciones = adquisiciones + inicial_euro
        num_adquisiciones += 1
        final_euro = final_dolar * final_cambio
        transmisiones = transmisiones + final_euro
        num_transmissiones += 1

        # aqui monto el diccionario con los resultados para cada trader copiado
        if e[2] is None:
            trader = 'Yo'
        else:
            trader = e[2]
        if trader not in estructura:
            estructura[trader]['profit'] = round(float(e[8].replace(',', '.')), 2)
            estructura[trader]['spread'] = round(float(e[7].replace(',', '.')), 2)
            estructura[trader]['fees'] = round(float(e[13].replace(',', '.')), 2)
            estructura[trader]['transacciones'] = 1
        else:
            estructura[trader]['profit'] += round(float(e[8].replace(',', '.')), 2)
            estructura[trader]['spread'] += round(float(e[7].replace(',', '.')), 2)
            estructura[trader]['fees'] += round(float(e[13].replace(',', '.')), 2)
            estructura[trader]['transacciones'] += 1

    # creo la lista de los trader copiados por orden alfabetico
    copiados = []
    for key in estructura:
        copiados.append(key)
    copiados = sorted(copiados, key=str.casefold)

    # proceso las transacciones realizadas
    sheet_2 = workbook['Transactions Report']
    equity_change_cerradas = 0
    equity_change_abiertas = 0
    fondos_aportados = 0
    gastos_deducibles_dolar = 0
    gastos_deducibles_euro = 0
    tipos_de_gastos_deducibles = ['Rollover Fee']

    for e in sheet_2.iter_rows(min_row=2, max_col=9, values_only=True):
        if e[2] in tipos_de_gastos_deducibles:
            fecha = datetime.strptime(e[0].strip()[:10], "%Y-%m-%d")
            for cambio in cambio_euro_dolar:
                if fecha >= cambio[0]:
                    cambio = cambio[1]
                    break
            gastos_deducibles_dolar = gastos_deducibles_dolar + e[5]
            gastos_deducibles_euro = gastos_deducibles_euro + e[5] * cambio

        if e[4] in ID_operaciones_cerradas:
            equity_change_cerradas += e[6]
        else:
            if e[2] == 'Deposit':
                fondos_aportados += e[5]
            else:
                equity_change_abiertas += e[6]

    # inicio la impresion de resultados
    posicion_fecha_cierre_primera_operacion = 'K' + str(sheet_1.max_row)
    posicion_fecha_apertura_primera_operacion = 'J' + str(sheet_1.max_row)
    print()
    print('--- Desglose por trader')
    print()

    total_transacciones = 0
    total_profit = 0
    total_fees = 0
    total_spread = 0
    for e in copiados:
        print('Trader:', e)
        print(' Operaciones cerradas =', estructura[e]['transacciones'])
        print(' Profit =', '{:>8}'.format('{:.2f}'.format(estructura[e]['profit']) + '$'))
        print(' Fees   =', '{:>8}'.format('{:.2f}'.format(estructura[e]['fees']) + '$'))
        print(' Neto   =', '{:>8}'.format('{:.2f}'.format((estructura[e]['profit'] + estructura[e]['fees'])) + '$'))
        print(' Spread =', '{:>8}'.format('{:.2f}'.format(estructura[e]['spread'] * (-1)) + '$'))
        total_transacciones = total_transacciones + estructura[e]['transacciones']
        total_profit += round(estructura[e]['profit'], 2)
        total_fees += round(estructura[e]['fees'], 2)
        total_spread += round(estructura[e]['spread'], 2)
    print()
    print('--- Resumen operaciones')
    print()
    print('Fondos aportados: ', '{:>8}'.format('{:.2f}'.format(fondos_aportados) + '$'))
    print('Operaciones cerradas =', total_transacciones)
    print('Fecha apertura/cierre primera operación:', sheet_1[posicion_fecha_apertura_primera_operacion].value, '-',
          sheet_1[posicion_fecha_cierre_primera_operacion].value)
    print('Fecha cierre última operación:                             ', sheet_1['K2'].value)
    print('Profit total   =', '{:>8}'.format('{:.2f}'.format(total_profit) + '$'))
    print('Fees totales   =', '{:>8}'.format('{:.2f}'.format(total_fees) + '$'))
    print('Neto total     =', '{:>8}'.format('{:.2f}'.format(total_profit + total_fees) + '$'))
    print('Spread total   =', '{:>8}'.format('{:.2f}'.format(total_spread * (-1)) + '$'))
    print()
    print('--- IRPF')
    print('Total adquisiciones =', '{:>8}'.format('{:.2f}'.format(adquisiciones) + '€'))
    print('Total transmisiones =', '{:>8}'.format('{:.2f}'.format(transmisiones) + '€'))
    print('Ganancia Patrimonial =', '{:>8}'.format('{:.2f}'.format(transmisiones - adquisiciones) + '€'))
    print('Gastos =', '{:>8}'.format('{:.2f}'.format(gastos_deducibles_euro)), '€ ', '(', format('{:.2f}'.format(gastos_deducibles_dolar)), '$)')
    print()
    input('Pulsa Enter para cerrar...')
